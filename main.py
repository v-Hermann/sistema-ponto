import os
import tkinter as tk
from datetime import datetime, timedelta
from tkinter import ttk
import openpyxl
from openpyxl import load_workbook

# Variável para adicionar dias para fins de teste (0 significa hoje)
time_delta = timedelta(days=0)

# Obtém o nome do mês atual para usar no nome do arquivo
mes_atual = (datetime.now() + time_delta).strftime('%B').lower()
meses_em_portugues = {
    'january': 'janeiro', 'february': 'fevereiro', 'march': 'marco',
    'april': 'abril', 'may': 'maio', 'june': 'junho',
    'july': 'julho', 'august': 'agosto', 'september': 'setembro',
    'october': 'outubro', 'november': 'novembro', 'december': 'dezembro'
}

def criar_planilha(filename):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Ponto"
    sheet['A1'] = "Data"
    sheet['B1'] = "Horário de Entrada"
    sheet['C1'] = "Horário de Saída"
    sheet['D1'] = "Horas Trabalhadas"
    sheet['E1'] = "Horas Extras"
    sheet['G1'] = "Total Horas Extras/Faltas do Mês"
    wb.save(filename)

def formatar_timedelta(tdelta):
    horas, restante = divmod(tdelta.seconds, 3600)
    minutos = restante // 60
    return f"{horas:02d}:{minutos:02d}"

def calcular_horas_trabalhadas(entrada, saida, carga_horaria):
    horas_trabalhadas = saida - entrada
    horas_extras = horas_trabalhadas - carga_horaria
    horas_trabalhadas_str = formatar_timedelta(horas_trabalhadas)
    horas_extras_str = formatar_timedelta(abs(horas_extras))

    if horas_extras < timedelta(0):
        horas_extras_str = f"-[{horas_extras_str}]"
    return horas_trabalhadas_str, horas_extras_str

def atualizar_total_horas_extras(sheet):
    total_segundos_extras = 0
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=5, max_col=5):
        cell = row[0].value
        if cell and cell != '00:00':
            # Remover colchetes e sinal negativo, se houver, para obter horas e minutos
            negativo = '-' in cell
            horas_minutos = cell.strip('[]-')
            horas, minutos = map(int, horas_minutos.split(':'))
            segundos = horas * 3600 + minutos * 60
            if negativo:
                total_segundos_extras -= segundos
            else:
                total_segundos_extras += segundos

    # Converter o total de segundos de volta para horas e minutos
    total_horas, total_minutos = divmod(abs(total_segundos_extras), 3600)
    total_minutos //= 60
    sinal = '-' if total_segundos_extras < 0 else ''
    sheet['G2'] = f"{sinal}{total_horas:02d}:{total_minutos:02d}"



def bater_ponto():
    agora = datetime.now() + time_delta
    data = agora.strftime('%d-%m-%Y')
    hora = agora.strftime('%H:%M')
    carga_horaria = timedelta(hours=6)  # Ajuste de acordo com a carga horária esperada

    wb = load_workbook(excel_filename)
    sheet = wb.active

    linha_data = None
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] == data:
            linha_data = row_index
            break

    mensagem = ""
    if linha_data is None:
        linha_data = sheet.max_row + 1
        sheet.cell(row=linha_data, column=1, value=data)
        sheet.cell(row=linha_data, column=2, value=hora)
        mensagem = "Primeiro ponto do dia registrado com sucesso."
        botao_ponto.config(state=tk.DISABLED)
    else:
        if sheet.cell(row=linha_data, column=3).value is None:
            sheet.cell(row=linha_data, column=3, value=hora)
            entrada = datetime.strptime(sheet.cell(row=linha_data, column=2).value, '%H:%M')
            saida = datetime.strptime(hora, '%H:%M')
            horas_trabalhadas_str, horas_extras_str = calcular_horas_trabalhadas(entrada, saida, carga_horaria)

            sheet.cell(row=linha_data, column=4, value=horas_trabalhadas_str)
            sheet.cell(row=linha_data, column=5, value=horas_extras_str)
            mensagem = f"Saída registrada.\nHorário de Entrada: {entrada.strftime('%H:%M')}\nHorário de Saída: {saida.strftime('%H:%M')}\nHoras Trabalhadas: {horas_trabalhadas_str}"
            botao_ponto.config(state=tk.DISABLED)  # Desabilita o botão também após a batida de saída
        else:
            mensagem = "Entrada e saída já registradas para hoje."
            botao_ponto.config(state=tk.DISABLED)

    atualizar_total_horas_extras(sheet)
    wb.save(excel_filename)
    wb.close()

    info_label.config(text=mensagem)


excel_filename = f'ponto_{meses_em_portugues[mes_atual]}.xlsx'
if not os.path.exists(excel_filename):
    criar_planilha(excel_filename)

janela = tk.Tk()
janela.title('Sistema de Ponto')
janela.geometry("400x200")
janela.configure(bg='#F0F0F0')

style = ttk.Style()
style.configure('TButton', background='#3498db', font=('Helvetica', 12, 'bold'), foreground='black')
style.map('TButton', background=[('active', '#2980b9')])

info_label = ttk.Label(janela, text="Bem-vinda ao Sistema de Ponto", font=('Helvetica', 12), background='#F0F0F0')
info_label.pack(pady=20)

botao_ponto = ttk.Button(janela, text='Bater Ponto', command=bater_ponto)
botao_ponto.pack()
botao_ponto.place(x=150, y=120)

janela.mainloop()
